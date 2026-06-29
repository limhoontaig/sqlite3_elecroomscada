# excel_report.py
import sqlite3
import shutil
import sys
import os
import zipfile
import re
from datetime import datetime, timedelta
import openpyxl
from db_manager import (DB_NAME, DATA_LABELS, METER_FIELDS, DB_DIR, 
                        create_manual_meter_table, get_field_inspections_for_date) # 👈 괄호로 묶고 이 함수를 추가해 줍니다.

TEMPLATE_NAME = "template_전기실_운영일지.xlsx"
TEMPLATE_IN_APPDATA = os.path.join(DB_DIR, TEMPLATE_NAME)

def get_bundle_template_path(relative_path):
    """
    PyInstaller (.exe) 환경과 일반 파이썬 (.py) 환경을 모두 지원하는 
    완벽한 템플릿 경로 추적 디버깅 함수입니다.
    """
    try:
        # 1. PyInstaller로 실행파일이 된 경우 (임시 폴더 경로 반환)
        base_path = sys._MEIPASS
        print(f"[DEBUG-EXE] 실행파일 환경 내부 경로 탐색: {base_path}")
        return os.path.join(base_path, relative_path)
    except Exception:
        # 2. 일반 파이썬 코드로 디버깅/실행 중인 경우
        # 💡 현재 이 코드 파일(excel_report.py)이 위치한 실제 절대 경로를 기준점으로 잡습니다!
        base_path = os.path.dirname(os.path.abspath(__file__))
        print(f"[DEBUG-DEV] 일반 파이썬 디버깅 환경 경로 탐색: {base_path}")
        return os.path.join(base_path, relative_path)

def ensure_excel_template():
    """앱데이터 폴더에 템플릿 파일이 없으면 자동으로 원본을 찾아서 복사해 둡니다."""
    print(f"[DEBUG] 최종 앱데이터 템플릿 타깃 경로: {TEMPLATE_IN_APPDATA}")
    
    if not os.path.exists(TEMPLATE_IN_APPDATA):
        print("⚠️ [DEBUG] 앱데이터 폴더 내에 템플릿 파일이 존재하지 않습니다. 복사를 시작합니다.")
        
        # 위에서 보정한 안전한 경로 함수 호출
        bundled_template = get_bundle_template_path(TEMPLATE_NAME)
        print(f"[DEBUG] 복사 소스가 될 원본 템플릿 예상 경로: {bundled_template}")
        
        if os.path.exists(bundled_template):
            # 앱데이터 폴더가 없으면 자동 생성
            if not os.path.exists(DB_DIR):
                os.makedirs(DB_DIR)
                print(f"[DEBUG] 앱데이터 디렉토리 생성 완료: {DB_DIR}")
                
            shutil.copy(bundled_template, TEMPLATE_IN_APPDATA)
            print(f"✅ [성공] 엑셀 템플릿 파일이 앱데이터에 안전하게 복구되었습니다.")
        else:
            print(f"❌ [오류] 원본 템플릿 파일({TEMPLATE_NAME})을 지정된 경로에서 찾을 수 없습니다. 파일명을 다시 확인하세요.")
    else:
        print("ℹ️ [DEBUG] 앱데이터 폴더에 이미 템플릿이 안전하게 존재하므로 복사를 생략합니다.")

# 프로그램 시작 시점에 무조건 디버깅 체크 구동
ensure_excel_template()

def clean_external_links_physically(file_path):
    """
    오류 팝업창의 근본 원인이 되는 엑셀 내부의 외부 링크 파편(externalLinks)들을
    Zipfile 레이어에서 완벽하게 박멸하는 물리 세척기 함수입니다.
    """
    temp_file_path = file_path + ".tmp"
    try:
        with zipfile.ZipFile(file_path, 'r') as zin:
            with zipfile.ZipFile(temp_file_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    filename = item.filename
                    if "externalLinks/" in filename:
                        continue
                    if filename == "xl/_rels/workbook.xml.rels":
                        data = zin.read(filename).decode('utf-8')
                        data_clean = re.sub(r'<Relationship[^>]*Type="[^"]*externalLink"[^>]*/>', '', data)
                        zout.writestr(filename, data_clean.encode('utf-8'))
                        continue
                    if filename == "xl/workbook.xml":
                        data = zin.read(filename).decode('utf-8')
                        data_clean = re.sub(r'<externalReferences>.*?</externalReferences>', '', data, flags=re.DOTALL)
                        zout.writestr(filename, data_clean.encode('utf-8'))
                        continue
                    zout.writestr(item, zin.read(filename))
        os.remove(file_path)
        os.rename(temp_file_path, file_path)
    except Exception as e:
        if os.path.exists(temp_file_path): os.remove(temp_file_path)
        print(f"[경고] 외부 링크 물리 청소 중 예외 발생: {e}")

def generate_excel_report(selected_date, target_dir=None):

    """
    21행의 일간 전력량 서식(min/max)과 22행의 월간 전력량 서식(start/end)을 
    DB 실측 통계 데이터로 매핑하고, 수식이 정상 작동하도록 주입하는 핵심 엔진입니다.
    """
    # ⭕ 변경 코드: 무조건 안전한 로컬 앱데이터 폴더의 템플릿을 소스로 사용합니다.
    template_file = TEMPLATE_IN_APPDATA

    # 🛠️ 수정 배치: target_dir이 지정되면 해당 폴더에, 없으면 기존처럼 소스 폴더에 저장
    if target_dir:
        output_file = os.path.join(target_dir, f"{selected_date}_전기실_운영일지.xlsx")
    else:  
        output_file = os.path.join(DB_DIR, f"{selected_date}_전기실_운영일지.xlsx")
    
    if not os.path.exists(TEMPLATE_IN_APPDATA):
        raise FileNotFoundError(f"템플릿 파일 [{template_file}]이 경로에 존재하지 않습니다.")
    
    # 1. 템플릿 안전 복사 및 워크북 로드
    shutil.copy(template_file, output_file)
    wb = openpyxl.load_workbook(output_file, data_only=False)
    
    # 오픈픽셀 자체 유령 링크 버그 예방책 마련
    if hasattr(wb, 'external_link_refs'): wb.external_link_refs = []
    if hasattr(wb, '_external_links'): wb._external_links = []

    dt = datetime.strptime(selected_date, "%Y-%m-%d")
    date_str_formatted = f"{dt.year}년 {dt.month:02d}월 {dt.day:02d}일"
    
    # 시트 명칭 동적 보정 및 할당 ("전력설비_운전현황" 또는 "전력설비_운영현황" 지원)
    ws_summary = None
    ws_detail = None
    for sheet in wb.worksheets:
        if "운전현황" in sheet.title or "운영현황" in sheet.title:
            ws_summary = sheet
            sheet.title = f"{selected_date}_전력설비_운전현황"
        elif "상세내역" in sheet.title:
            ws_detail = sheet
            sheet.title = f"{selected_date}_상세내역"
            
    if not ws_summary: ws_summary = wb.worksheets[0]
    if not ws_detail: ws_detail = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # =========================================================================
    # [파트 1] 전력설비 운영 현황 분석 자료 - 최고/최저값 덮어쓰기 (상단 구역)
    # =========================================================================
    c.execute('SELECT * FROM daily_extremes WHERE log_date = ?', (selected_date,))
    rows_extreme = c.fetchall()
    c.execute('PRAGMA table_info(daily_extremes)')
    columns_extreme = [col[1] for col in c.fetchall()]
    
    extremes_dict = {}
    for r in rows_extreme:
        ext_type = r[columns_extreme.index('extreme_type')]
        for col_name in DATA_LABELS:
            if col_name in columns_extreme:
                val = r[columns_extreme.index(col_name)]
                extremes_dict[(ext_type, col_name)] = round(val, 1) if val is not None else "-"

    for row in ws_summary.iter_rows(max_row=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip().replace(" ", "")
                if "일시:" in val_strip or ("일시" in val_strip and ":" in val_strip):
                    cell.value = f"일시 : {date_str_formatted}"
                elif "max(" in val_strip:
                    label = cell.value.split('"')[1].strip() if '"' in cell.value else cell.value.replace("max(", "").replace(")", "").strip()
                    cell.value = extremes_dict.get(("MAX", label), "-")
                elif "min(" in val_strip:
                    label = cell.value.split('"')[1].strip() if '"' in cell.value else cell.value.replace("min(", "").replace(")", "").strip()
                    cell.value = extremes_dict.get(("MIN", label), "-")

    # =========================================================================
    # 💡 [파트 2 완벽 반영] 2. 사용 전력량 현황 자동 데이터 치환 (21행 및 22행 영역)
    # =========================================================================
    day_min_val = "-"
    day_max_val = "-"
    if "KEP_P_kWh" in columns_extreme:
        day_min_val = extremes_dict.get(("MIN", "KEP_P_kWh"), "-")
        day_max_val = extremes_dict.get(("MAX", "KEP_P_kWh"), "-")
    
    if day_min_val == "-" or day_max_val == "-":
        c.execute('SELECT MIN(KEP_P_kWh), MAX(KEP_P_kWh) FROM raw_data WHERE log_date = ?', (selected_date,))
        res_raw = c.fetchone()
        if res_raw:
            if day_min_val == "-" and res_raw[0] is not None: day_min_val = round(res_raw[0], 1)
            if day_max_val == "-" and res_raw[1] is not None: day_max_val = round(res_raw[1], 1)

    start_of_month = f"{dt.year}-{dt.month:02d}-01"
    c.execute('''
        SELECT KEP_P_kWh FROM raw_data 
        WHERE log_date >= ? AND log_date <= ? AND KEP_P_kWh IS NOT NULL 
        ORDER BY log_date ASC, log_time ASC LIMIT 1
    ''', (start_of_month, selected_date))
    month_start_res = c.fetchone()
    month_start_val = round(month_start_res[0], 1) if month_start_res and month_start_res[0] is not None else "-"

    for row_idx in [21, 22]:
        for cell in ws_summary[row_idx]:
            if cell.value and isinstance(cell.value, str):
                text_clean = cell.value.strip().replace(" ", "")
                if row_idx == 21:
                    if 'min("KEP_P_kWh")' in text_clean: cell.value = day_min_val
                    elif 'max("KEP_P_kWh")' in text_clean: cell.value = day_max_val
                    elif 'E21-C21' in text_clean: cell.value = "=E21-C21"
                elif row_idx == 22:
                    if 'start("KEP_P_kWh")' in text_clean: cell.value = month_start_val
                    elif 'end("KEP_P_kWh")' in text_clean: cell.value = day_max_val
                    elif 'E22-C22' in text_clean: cell.value = "=E22-C22"

    # =========================================================================
    # [파트 3] 한전 메인 적산 사용량 하단 요약 테이블 연동 구역 (26행 이하)
    # =========================================================================
    today_mwh = day_max_val if day_max_val != "-" else "-"
    prev_mwh = day_min_val if day_min_val != "-" else "-"
    diff_mwh = round(today_mwh - prev_mwh, 1) if today_mwh != "-" and prev_mwh != "-" else "-"

    for row in ws_summary.iter_rows(min_row=26, max_row=30): # 루프 범위 안전 제한
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip()
                if val_strip == "1469.79": cell.value = today_mwh
                elif val_strip == "1464.06": cell.value = prev_mwh
                elif val_strip == "5.7300000000000182": cell.value = diff_mwh

    # =========================================================================
    # 💡 [최종 수정본] 수동 검침(manual_meter_logs) 안전 데이터 연동 구역 (C27 ~ J35)
    # =========================================================================
    create_manual_meter_table()
    
    # 전일 날짜 계산
    prev_day = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # 전월 마지막 날 계산 (고정)
    first_day_of_current_month = dt.replace(day=1)
    last_day_of_prev_month_dt = first_day_of_current_month - timedelta(days=1)
    prev_month_last_day = last_day_of_prev_month_dt.strftime('%Y-%m-%d')
    
    print(f"[DEBUG-METER] 금일: {selected_date}, 전일: {prev_day}, 전월마지막날: {prev_month_last_day}")

    # 금일, 전일, 전월 데이터 각각 독립적으로 로드 (서로 에러 간섭 없도록 분리)
    meter_data = {"금일": {}, "전일": {}, "전월": {}}
    
    for label, target_d in [("금일", selected_date), ("전일", prev_day), ("전월", prev_month_last_day)]:
        fields_str = ", ".join(METER_FIELDS)
        c.execute(f'SELECT {fields_str} FROM manual_meter_logs WHERE log_date = ?', (target_d,))
        res = c.fetchone()
        
        if res:
            meter_data[label] = {field: res[idx] for idx, field in enumerate(METER_FIELDS)}
            print(f"[DEBUG-METER] ✅ {label}({target_d}) 데이터 로드 성공")
        else:
            meter_data[label] = {field: None for field in METER_FIELDS}
            print(f"[DEBUG-METER] ⚠️ {label}({target_d}) 데이터가 DB에 없습니다. (공백 처리)")

    # C27부터 J35 영역 스캔 및 안전하게 치환
    for row in ws_summary.iter_rows(min_row=27, max_row=35, min_col=3, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip().replace(" ", "")
                
                match = re.match(r'(금일|전일|전월)\(["\']?([a-zA-Z0-9_]+)["\']?\)', val_strip)
                if match:
                    time_type = match.group(1)   # 금일 / 전일 / 전월
                    field_name = match.group(2)  # main_active 등
                    
                    if field_name in METER_FIELDS:
                        # 🔥 안전 무력화 방어 조치: 해당 구조나 필드가 비어있어도 프로그램이 튕기거나 멈추지 않습니다.
                        target_dict = meter_data.get(time_type, {})
                        db_val = target_dict.get(field_name, None)
                        
                        # 데이터가 있으면 실수형 숫자로 채우고, 없으면 빈칸("") 처리하여 다음 셀로 넘어감
                        cell.value = float(db_val) if db_val is not None else ""

    # =================================================================
    # [파트 3-4 신규 추가] 하드코딩 탈피: E40:G48 셀을 순회하며 점검 데이터 자동 매핑
    # =================================================================
    try:
        print(f"[DEBUG] 현장 점검 데이터 엑셀 매핑 시작 (날짜: {selected_date})")
        
        # 1. 상단에서 임포트한 함수를 통해 오늘 날짜의 1,2,3차 점검자/시간 딕셔너리 획득
        inspection_data = get_field_inspections_for_date(selected_date)
        
        # 2. 과장님이 제안하신 E40부터 G48까지의 범위를 순회 (행과 열을 돌며 탐색)
        for row in ws_summary["E40:G48"]:
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 공백을 제거하고 순수 텍스트만 비교합니다.
                    cell_text = cell.value.strip()
                    
                    # --- 1차 점검 매핑 ---
                    if "1차" in cell_text and "점검자" in cell_text:
                        cell.value = inspection_data[1]["name"] if inspection_data[1]["name"] else "-"
                    elif "1차" in cell_text and "시간" in cell_text:
                        cell.value = inspection_data[1]["time"] if inspection_data[1]["time"] else "-"
                        
                    # --- 2차 점검 매핑 ---
                    elif "2차" in cell_text and "점검자" in cell_text:
                        cell.value = inspection_data[2]["name"] if inspection_data[2]["name"] else "-"
                    elif "2차" in cell_text and "시간" in cell_text:
                        cell.value = inspection_data[2]["time"] if inspection_data[2]["time"] else "-"
                        
                    # --- 3차 점검 매핑 ---
                    elif "3차" in cell_text and "점검자" in cell_text:
                        cell.value = inspection_data[3]["name"] if inspection_data[3]["name"] else "-"
                    elif "3차" in cell_text and "시간" in cell_text:
                        cell.value = inspection_data[3]["time"] if inspection_data[3]["time"] else "-"
                        
        print("[DEBUG] 현장 점검 데이터 엑셀 유연 매핑 완료")
        
    except Exception as e:
        # 혹시 모를 에러 발생 시 프로그램 전체가 멈추지 않고 로그만 남기도록 유도
        print(f"[ERROR] 현장 점검 서식 순회 매핑 중 오류 발생: {e}")
    # =================================================================

    # =========================================================================
    # [파트 4] 상세내역 시트 - 시간별 데이터 완벽 주입 (상/하단 24시간 입력 제어)
    # =========================================================================
    for row in ws_detail.iter_rows(max_row=3):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "일자" in cell.value:
                cell.value = f"일자 : {date_str_formatted}"

    avg_selects = ", ".join([f'AVG("{name}")' for name in DATA_LABELS])
    c.execute(f'SELECT CAST(strftime("%H", log_time) AS INTEGER) as hour, {avg_selects} FROM raw_data WHERE log_date = ? GROUP BY hour', (selected_date,))
    hourly_rows = c.fetchall()
    
    hourly_data = {}
    for r in hourly_rows:
        hour = r[0]
        for idx, label in enumerate(DATA_LABELS):
            if r[idx + 1] is not None:
                hourly_data[(hour, label)] = round(r[idx + 1], 1)

    for row_idx, row in enumerate(ws_detail.iter_rows(min_row=4, max_row=60), start=4):
        first_cell_val = str(row[0].value).strip() if row[0].value is not None else ""
        if first_cell_val == "0":
            col_mapping = {}
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str):
                    clean_label = cell.value.strip().replace('"', '')
                    if clean_label in DATA_LABELS:
                        col_mapping[col_idx] = clean_label
            
            if col_mapping:
                for h in range(24):
                    target_row = row_idx + h
                    ws_detail.cell(row=target_row, column=1).value = h
                    for col_idx, label in col_mapping.items():
                        val = hourly_data.get((h, label), "-")
                        ws_detail.cell(row=target_row, column=col_idx + 1).value = val

    conn.close()
    wb.save(output_file)
    
    # 엑셀 파일 내부 외부 수식 관계 물리적 청소 기능 호출
    clean_external_links_physically(output_file)
    print(f"[업데이트 완료] 전력량 및 수동 검침 연동 성공: {output_file}")

if __name__ == "__main__":
    generate_excel_report("2026-05-21")